from selenium import webdriver
from openpyxl import load_workbook

# 엑셀 오픈
wb = load_workbook("노래.xlsx")
# 엑셀 시트 오픈
ws = wb["노래검색"]
# A열 값 다들고옴
col = ws["A"]

# A열값 하나하나씩
for cell in col:
    name= cell.value
    # 값 인터넷 검색할수있게 수정
    name.encode('utf-8')
    # 크롤링하다 오류가나면 종료
    try:
        driver = webdriver.Chrome("./chromedriver")
        # 네이버 검색주소
        driver.get("https://search.naver.com/search.naver?where=nexearch&sm=top_hty&fbm=1&ie=utf8&query=%s" % name)
        # 가수이름
        elem =driver.find_element_by_xpath("//*[@id='main_pack']/div[1]/div[1]/div[1]/div/span[1]/a")
        # 가수이름 값의 행 B열에 삽입
        ws["B%s" % cell.row]=elem.text
        elem = driver.find_element_by_xpath("//*[@id='main_pack']/div[1]/div[1]/div[1]/div/span[3]")
        # 노래연도 값의 행 C열에 삽입 하지만 .은 -로 변환
        ws["C%s" % cell.row]=elem.text.replace(".","-")
        # 크롬 종료
        driver.quit()
    except Exception as e:
        print(e)
    finally:
        driver.quit()

# 마지막으로 완료된 엑셀파일 저장
wb.save("제작완료.xlsx")